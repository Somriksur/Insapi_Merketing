"""Excel/CSV import for the Credits ledger."""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


VALID_KINDS = {"advance", "credit", "refund", "adjustment"}

COLUMN_ALIASES = {
    "client_name": ["client_name", "client", "name", "customer", "party"],
    "client_id": ["client_id", "id"],
    "kind": ["kind", "type", "entry_type"],
    "amount": ["amount", "value", "₹", "rupees", "inr"],
    "note": ["note", "notes", "remark", "remarks", "description", "comment"],
    "date": ["date", "txn_date", "entry_date"],
}


def _norm(s: Any) -> str:
    return str(s or "").strip().lower().replace(" ", "_")


def _resolve_columns(cols: List[str]) -> Dict[str, str]:
    """Map a canonical key -> actual column name from the sheet."""
    n_to_real = {_norm(c): c for c in cols}
    resolved = {}
    for canon, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in n_to_real:
                resolved[canon] = n_to_real[alias]
                break
    return resolved


def _parse_date(v: Any) -> Optional[str]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date().isoformat()
            except Exception:
                pass
    try:
        return pd.to_datetime(v).date().isoformat()
    except Exception:
        return None


async def parse_and_import(
    db,
    data: bytes,
    filename: str,
    auto_create_clients: bool = True,
) -> Dict[str, Any]:
    """Parse an Excel/CSV upload and add rows to the credits collection.

    Returns a dict: {imported, skipped, created_clients, errors:[{row, reason}]}.
    """
    lname = (filename or "").lower()
    try:
        if lname.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(data))
        else:
            df = pd.read_excel(io.BytesIO(data), engine="openpyxl")
    except Exception as e:
        return {"imported": 0, "skipped": 0, "created_clients": 0, "errors": [{"row": 0, "reason": f"Could not read file: {e}"}]}

    if df.empty:
        return {"imported": 0, "skipped": 0, "created_clients": 0, "errors": [{"row": 0, "reason": "Sheet is empty"}]}

    cols = _resolve_columns(list(df.columns))
    if "amount" not in cols:
        return {"imported": 0, "skipped": 0, "created_clients": 0, "errors": [
            {"row": 0, "reason": "Missing required column 'amount'. Allowed headers: client_name, kind, amount, note, date"}
        ]}
    if "client_name" not in cols and "client_id" not in cols:
        return {"imported": 0, "skipped": 0, "created_clients": 0, "errors": [
            {"row": 0, "reason": "Missing 'client_name' (or 'client_id') column"}
        ]}

    # Cache clients (name -> id)
    existing = await db.clients.find({}, {"_id": 0}).to_list(5000)
    name_to_id = {(c.get("name") or "").strip().lower(): c["id"] for c in existing}
    id_set = {c["id"] for c in existing}

    imported = 0
    skipped = 0
    created_clients = 0
    errors: List[Dict[str, Any]] = []
    new_credits: List[Dict[str, Any]] = []
    new_clients: List[Dict[str, Any]] = []

    for idx, row in df.iterrows():
        row_no = int(idx) + 2  # account for header line
        try:
            amt_raw = row.get(cols["amount"])
            if pd.isna(amt_raw) or amt_raw is None or amt_raw == "":
                skipped += 1
                continue
            amount = float(amt_raw)
            if amount == 0:
                skipped += 1
                continue

            client_id = None
            if "client_id" in cols:
                cid = str(row.get(cols["client_id"], "") or "").strip()
                if cid and cid in id_set:
                    client_id = cid

            if not client_id and "client_name" in cols:
                cname = str(row.get(cols["client_name"], "") or "").strip()
                if not cname:
                    errors.append({"row": row_no, "reason": "missing client name"})
                    skipped += 1
                    continue
                key = cname.lower()
                if key in name_to_id:
                    client_id = name_to_id[key]
                elif auto_create_clients:
                    import uuid
                    new_id = str(uuid.uuid4())
                    new_clients.append({
                        "id": new_id,
                        "name": cname,
                        "company": "",
                        "email": "",
                        "whatsapp": "",
                        "address": "",
                        "rating": 5,
                        "notes": "Imported from Excel",
                        "tags": ["imported"],
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    })
                    name_to_id[key] = new_id
                    id_set.add(new_id)
                    client_id = new_id
                    created_clients += 1
                else:
                    errors.append({"row": row_no, "reason": f"client '{cname}' not found"})
                    skipped += 1
                    continue

            kind = "credit"
            if "kind" in cols:
                k = _norm(row.get(cols["kind"]))
                if k in VALID_KINDS:
                    kind = k
                elif k:
                    errors.append({"row": row_no, "reason": f"unknown kind '{k}', using 'credit'"})

            date_iso = None
            if "date" in cols:
                date_iso = _parse_date(row.get(cols["date"]))

            note = ""
            if "note" in cols:
                v = row.get(cols["note"])
                if v is not None and not (isinstance(v, float) and pd.isna(v)):
                    note = str(v).strip()

            import uuid
            entry = {
                "id": str(uuid.uuid4()),
                "client_id": client_id,
                "kind": kind,
                "amount": amount,
                "note": note,
                "date": date_iso,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            new_credits.append(entry)
            imported += 1
        except Exception as e:
            errors.append({"row": row_no, "reason": str(e)})
            skipped += 1

    if new_clients:
        await db.clients.insert_many(new_clients)
    if new_credits:
        await db.credits.insert_many(new_credits)

    return {
        "imported": imported,
        "skipped": skipped,
        "created_clients": created_clients,
        "errors": errors[:50],
    }


def build_template_xlsx() -> bytes:
    """Return a downloadable Excel template with headers + example rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Credits"

    headers = ["client_name", "kind", "amount", "date", "note"]
    header_fill = PatternFill("solid", fgColor="0B0B0B")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 40

    # example rows
    examples = [
        ["Aarav Mehta", "advance", 5000, "2026-05-08", "Advance for May reel batch"],
        ["Neha Kapoor", "refund", 1500, "2026-05-09", "Over-charged for second revision"],
        ["Rahul Shah", "credit", 750, "2026-05-10", "Loyalty discount"],
    ]
    for r, row in enumerate(examples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    # legend sheet
    legend = wb.create_sheet("Notes")
    legend.append(["Column", "Required", "Allowed values / format", "Example"])
    legend.column_dimensions["A"].width = 16
    legend.column_dimensions["B"].width = 10
    legend.column_dimensions["C"].width = 50
    legend.column_dimensions["D"].width = 30
    legend.append(["client_name", "yes", "Matches existing client (case-insensitive). New ones are auto-created.", "Aarav Mehta"])
    legend.append(["kind", "no", "advance | credit | refund | adjustment (default: credit)", "advance"])
    legend.append(["amount", "yes", "Positive number in INR", "1500"])
    legend.append(["date", "no", "YYYY-MM-DD or DD/MM/YYYY", "2026-05-08"])
    legend.append(["note", "no", "Free text", "Advance for next batch"])
    for cell in legend[1]:
        cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
